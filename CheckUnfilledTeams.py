import datetime
import os
import win32com.client as win32
from datetime import datetime
import calendar
import tempfile
import subprocess

from PySide6.QtCore import QDateTime
import openpyxl
import openpyxl.cell
import openpyxl.workbook
import openpyxl.worksheet
import openpyxl.worksheet.worksheet

from copyPasteLinksOfPDF import copyPasteLinksofPDF
from LogData import Database
from applicationData import dataBaseColumnsAndDataTypes,logFilePath,reportMonth,reportWeek,settingsSaveFile


supportedExcelExtensions=[".xlsx",".xlsm",".xltx",".xltm"]

#reportVerifier class contains information about report, methods to check and generate reports
class KPIreportVerifier:
    def __init__(self,reportName:str,checkingFrequency:str="",reportPDFLocation:str=""
                 ,reportTemplatePDFLocation:str=""):
        self.report_location:str
        self.teams = []
        self.unfilled_teams = []
        self.responsibleData:dict[str,dict[str,list[str]]]={}
        self.reportName:str=reportName
        self.isEveryoneFilled=False
        self.isExcelPathPresent=False
        self.isReportGenerated=False
        self.isSuccessfullyCheckedWithoutErrors=False
        self.checkingFrequency=checkingFrequency
        self.reportPDFName:str
        self.reportPDFLocation:str=reportPDFLocation
        self.reportTemplatePDFLocation:str=reportTemplatePDFLocation
        self.MacroModule:str
        self.macroName:str
        self.isExternalDataRefreshRequired:bool=False
        self.tempReportPath:str=""
        self.checkingIterationsRan:int=1
        
        self.getReportGeneratedStatus()

    #If external data refresh is required win32.com used to open file in excel and refresh all queries and external data
    #leaves the report file unmodified when checking report by creating a temporary file 
    def refreshAndSaveInTempPath(self):
        if self.tempReportPath and os.path.exists(self.tempReportPath):
            try:
                os.remove(self.tempReportPath)
            except Exception as e:
                print(f"Error removing temp file: {self.tempReportPath}\n{e}")
        excel = win32.DispatchEx("Excel.Application")
        excel.DisplayAlerts=False
        excel.AskToUpdateLinks = False
        workbook = excel.Workbooks.Open(Filename=self.report_location,UpdateLinks=True)
        excel.CalculateUntilAsyncQueriesDone()
        tempFileName = f"temp_{self.reportName}_{datetime.now().strftime('%d-%b-%y_%I-%M_%p')}.xlsm"
        self.tempReportPath =tempfile.gettempdir()+"\\"+tempFileName
        try:
            workbook.SaveAs(Filename=self.tempReportPath, FileFormat=52)  
        except Exception as error:
            print(f"Error while saving {self.tempReportPath}\n{error}")
        finally:
            workbook.Close(SaveChanges=True)
            excel.Quit()
            workbook=  None
            excel = None
        
    #adds team after creating the class 
    def add_team(self, team_data,responsibleData:dict[str,list[str]]):
        self.teams.append(team_data)
        self.responsibleData[team_data]=responsibleData

    #can be used to get cells need to be filled if existing team is passed
    def getResponsibleData(self,team_data):
        return self.responsibleData[team_data]
    
    def getResponsibleSheets(self, team_data):
        return list(self.responsibleData[team_data].keys())
    
    def getResponsibleCells(self, team_data):
        return self.responsibleData[team_data]

    #report checking procedure
    def get_teams_with_unfilled_cells(self):
        self.unfilled_teams:list[str]=[]
        print(f"Currently checking {self.reportName}")
        print (f"No of times Rechecked : {self.checkingIterationsRan} ")
        self.isTeamAreDefined = bool(len(self.teams))
        self.isResponsibleDataPresent = bool(len(self.responsibleData))
        self.isExcelPathPresent=False
        if os.path.exists(self.report_location):
            self.isExcelPathPresent = any(self.report_location.endswith(ext) for ext in supportedExcelExtensions)

        # if any occurs when generating report all teams stored in the object is returned
        if not (self.isTeamAreDefined and self.isResponsibleDataPresent and self.isExcelPathPresent ):
            print(f"    Error While checking team\n"
                  f"    TeamsDeclared:{self.isTeamAreDefined}\n"
                  f"    ResponsibleDataDeclared:{self.isResponsibleDataPresent}\n"
                  f"    ExcelPathDeclared:{self.isExcelPathPresent}\n")
            return list(self.teams)

        if self.isExternalDataRefreshRequired:
            self.refreshAndSaveInTempPath()
            currentReportPath=self.tempReportPath
        else:
            currentReportPath=self.report_location

        workbook = openpyxl.load_workbook(currentReportPath,read_only=True,data_only=True)
        for team_data, sheet_cell_dict in self.responsibleData.items():
            for sheet_name, cell_list in sheet_cell_dict.items():
                if sheet_name in workbook.sheetnames:
                    worksheet:openpyxl.worksheet.worksheet.Worksheet = workbook[sheet_name]
                else:
                    print(f"Sheet {sheet_name} not found in the workbook.")
                    return list(self.teams)
                for cell_address in cell_list:
                    try:
                        cell_value: list[openpyxl.cell.Cell] = worksheet[cell_address].value 
                    except AttributeError as error:
                        self.unfilled_teams.append(team_data)
                        print(f"{team_data} Cell address:{cell_address} is not found in {sheet_name}")
                        break
                    if not (isinstance(cell_value,int) or isinstance(cell_value,float)):
                        self.unfilled_teams.append(team_data)
                        break
        workbook.close()
        if len(self.unfilled_teams)==0:
            self.isEveryoneFilled=True
        self.logToDatabase()
        self.isSuccessfullyCheckedWithoutErrors=True
        if self.isExternalDataRefreshRequired:
            os.remove(self.tempReportPath)
            self.tempReportPath=""
        return list(set(self.unfilled_teams))
    
    #gets isreportgenerated and recheck iterations ran from database for current week or month and stores in the object
    def getReportGeneratedStatus(self)->bool:
        logDatabase=Database(dataBasePath=fr"{logFilePath}",TableName=self.reportName,columnsAndDataTypes=dataBaseColumnsAndDataTypes)
        logDatabase.connection.commit()

        latestRow=logDatabase.getLatestRow(tableName=self.reportName)

        if latestRow==None or len(latestRow)==0:
            print (f"No Data found in the database table {self.reportName}")
            return False

        if len(dataBaseColumnsAndDataTypes.keys())!=len(latestRow):
            print("Database columns and application data doesn't match")
            print(f"Database columns : {logDatabase.getColumns(tableName=self.reportName)}")
            print(f"Application columns : {list(dataBaseColumnsAndDataTypes.keys())}")
            return False
        
        if latestRow[1]==str(calendar.month_name[reportMonth]) and latestRow[2]==str(reportWeek):
            if latestRow[5]=='True':
                self.isReportGenerated:bool=True
            else:
                self.isReportGenerated=False
            self.checkingIterationsRan=int(latestRow[7])
        else:
            self.isReportGenerated:bool=False
            self.checkingIterationsRan=1
        logDatabase.disconnect(saveChanges=False)
        return self.isReportGenerated
    
    #Logs necessary information to database for future use
    #Takes log when get_teams_with_unfilled_cells runned successfully without errors
    def logToDatabase(self):  
        logDatabase=Database(dataBasePath=fr"{logFilePath}")
        logData:dict={}
        if not len(self.unfilled_teams)==0:
            currentTime=QDateTime().currentDateTime()
            if self.checkingFrequency=="Weekly":
                recheckingWaitIntervalHours:float=settingsSaveFile.get("Weekly_Rechecking_Frequency", 1)
            else:
                recheckingWaitIntervalHours:float=settingsSaveFile.get("Monthly_Rechecking_Frequency", 3)
            recheckingTime=currentTime.addSecs(int(recheckingWaitIntervalHours*60*60))
            recheckingTime=recheckingTime.toString("dd-MM-yyyy hh:mm:ss")
        else:
            recheckingTime="-"
        column=list(dataBaseColumnsAndDataTypes.keys()) 
        logData[column[1]]=calendar.month_name[reportMonth]
        logData[column[2]]=reportWeek
        logData[column[3]]=str(self.isEveryoneFilled)
        logData[column[4]]=str(self.unfilled_teams)
        logData[column[5]]=str(self.isReportGenerated)
        logData[column[6]]=QDateTime.currentDateTime().toString("dd-MM-yyyy hh:mm:ss")
        logData[column[8]]=recheckingTime
        #checking for iterations ran data if not found taking 1, if found taking from database and incrementing by 1
        try:
            tempIterationsRan=int(logDatabase.getLatestData(tableName=self.reportName, columnName=column[7]))
            lastRecordedMonth=logDatabase.getLatestData(tableName=self.reportName, columnName=column[1])
            lastRecordedWeek=int(logDatabase.getLatestData(tableName=self.reportName, columnName=column[2]))
        except ValueError:
            tempIterationsRan=None
            lastRecordedMonth=None
            lastRecordedWeek=None
        if isinstance(tempIterationsRan,int)and calendar.month_name[reportMonth]==lastRecordedMonth and reportWeek==lastRecordedWeek:
            logData[column[7]]=tempIterationsRan+1
            self.checkingIterationsRan=tempIterationsRan+1
        else:
            logData[column[7]]=1
        logDatabase.insertData(tableName=self.reportName,columnAndValue=logData,saveChanges=True)
        logDatabase.disconnect(saveChanges=True)

    #default generate report procedure runs a macro,copy internal links from template to generated pdf
    #mark report generation status in the database
    def generateReport(self) -> None:        
        
        if self.isEveryoneFilled==False:
            print("Report completion check failed")

        isOnedriveClosed=stopOneDrive()
        didMacroRun=runExcelMacro(excelFilePath=self.report_location, 
                                  modulename=self.MacroModule, macroName=self.macroName,
                                  saveExcelFile= self.isExternalDataRefreshRequired==True)
        isLinksCopied=copyPasteLinksofPDF(sourcePDF=self.reportTemplatePDFLocation,destinationPDF=self.reportPDFLocation)
        if not (didMacroRun and isLinksCopied):
            print(f"Report generation failed"
                    f"Macro ran ?: {didMacroRun}"
                    f"links copied ?: {isLinksCopied}")
        if didMacroRun and isLinksCopied:
            self.isReportGenerated=True
        else:
            self.isReportGenerated=False
        if isOnedriveClosed:
            startOneDrive()
        self.markReportGenerationStatus()
    
    #update the current iteration data if report generation is completed
    def markReportGenerationStatus(self):
        column=list(dataBaseColumnsAndDataTypes.keys())
        logDatabase=Database(dataBasePath=fr"{logFilePath}")
        latestData=logDatabase.getLatestRow(tableName=self.reportName)
        reportMonthName=str(calendar.month_name[reportMonth])
        if reportMonthName in latestData and str(reportWeek) in latestData:
            logDatabase.changeLatestData(tableName=self.reportName,columnName=column[5],value=str(self.isReportGenerated))
            logDatabase.changeLatestData(tableName=self.reportName,columnName=column[8],value="-")
        else:
            print(f"markReportGenerationStatus failed cannot find the report for month {reportMonth} and week {reportWeek}")
        logDatabase.disconnect(saveChanges=True)


#not used
def readExcelCell(excelFilePath: str, sheetName: str, cellAddress: str):
    if not os.path.exists(excelFilePath):
        print("Excel file not found")
        return None
        
    workbook = openpyxl.load_workbook(excelFilePath, read_only=True, data_only=True)
    if sheetName not in workbook.sheetnames:
        print(f"Sheet {sheetName} not found in the workbook.")
        return None
    worksheet = workbook[sheetName]
    try:
        cell_value = worksheet[cellAddress].value
    except AttributeError as error:
        print(f"Cell address:{cellAddress} is not found in {sheetName}")
        return None
    finally:
        if 'workbook' in locals():
            workbook.close()
        worksheet = None
        workbook = None
    return cell_value

def startOneDrive():
    oneDriveExePaths=[
        os.path.join(os.environ["LOCALAPPDATA"], "Microsoft", "OneDrive","OneDrive.exe" ),
        os.path.join(os.environ["PROGRAMFILES"], "Microsoft OneDrive","OneDrive.exe" ),
        os.path.join(os.environ["PROGRAMFILES(X86)"], "Microsoft OneDrive","OneDrive.exe"),
        os.path.join(os.environ["USERPROFILE"], "OneDrive")
    ]
    for path in oneDriveExePaths:
        if os.path.exists(path):
            try:
                subprocess.Popen(executable=path,args="/background")
            except subprocess.CalledProcessError:
                print(f"Failed to start OneDrive at {path}")
            finally:
                return
    

def stopOneDrive()->bool:

    outputCheck=subprocess.check_output('tasklist /fi "imagename eq OneDrive.exe"', shell=True)
    if "No tasks are running" in str(outputCheck) : 
        return False
    
    try:
        subprocess.run(["taskkill", "/f", "/im", "OneDrive.exe"], shell=True)
    except subprocess.CalledProcessError:
        print(f"onedrive not running or not installed.")
        return False
    return True
    
    

#runs a macro using win32com returns true if macro ran successfully
def runExcelMacro(excelFilePath:str,modulename:str,macroName:str,saveExcelFile:bool)-> bool:
        if os.path.exists(excelFilePath)==False:
            print("Excel file not found")
            return False
        if (modulename=="" or macroName==""):
            print("Macro not defined")
            return False
        excelApp=win32.DispatchEx("Excel.Application")
        excelApp.Visible=False
        workBook=excelApp.Workbooks.Open(Filename=excelFilePath,UpdateLinks=True)
        workBook.RefreshAll()
        excelApp.CalculateUntilAsyncQueriesDone()
        fileName=excelFilePath.split(r"/")[-1]
        try:
            excelApp.Application.Run(f"'{fileName}'!{modulename}.{macroName}")
            print(f"successfully ran macro {macroName}")
        except Exception as e:
            print(f"Error running macro: {e}")
            return False
        finally:
            workBook.Close(SaveChanges=saveExcelFile)
            excelApp.Quit()
            workBook = None
            excelApp = None
            return True
            
        
    

if __name__ == "__main__":
    ...    

